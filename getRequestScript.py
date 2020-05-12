# -*- coding: utf-8 -*-
"""
Created on Mon Jan 13 16:25:17 2020

@author: T2207524
"""
import openpyxl as excel
import datetime
import urllib.request as request
import os

class openFile:
    dataList = []   # list to store websites for request
    def getSheet():
        i = 2           #data starts in the second row 
        workBook = excel.load_workbook(filename = r'Made up URI to protect security\sites_to_check_status.xlsx', read_only=True)
        workSheet = workBook['Sheet1'] # Specifing the name of the worksheet we are reading from
     
        #get data from excel spreadsheet cells
        cell = 'A' + str(i)     #cell location of data
        data = workSheet[cell].value
        while data != None:
            openFile.dataList.append(data)
            i = i + 1
            cell = 'A' + str(i)
            data = workSheet[cell].value
        else:
            workBook.close() 
            return openFile.dataList
        
class getRequest:
    resultsList = []  # where the get request will be stored. 
    def returnRequestStatus(List):
        for item in List:
            try:
                returnCode = request.urlopen(item).getcode() # doing the get request and returning the http response code.
                getRequest.resultsList.append(returnCode)
            except ValueError:
                #print("error " + item)
                getRequest.resultsList.append('Error with format of website in spreadsheet. Fix before next run') # will write to excel spreadsheet if URL is incorrect. 
        return getRequest.resultsList
            
class resultFile:
    
    def createFile():
        os.chdir('C:/Users/' +os.getlogin()+ '/Downloads/') # where we are saving the file name
        now = datetime.datetime.now()
        timeStamp = now.strftime('%Y-%m-%d, %H-%M')
        fileName = 'Website Status '+str(timeStamp)+'.xlsx' #created saved file name
        
        resultsWorkBook = excel.workbook.Workbook()         # creating workbook
        ws = resultsWorkBook.active                         # making the sheet active
        ws.title = 'Results Sheet'
        
        i = 1
        k = 1
        l = 1
        for x in openFile.dataList:                       # write website names to excel spreadsheet
            ws['A' + str(i)] = x
            i = i + 1
        for y in getRequest.resultsList:                    # write return codes to excel spreadsheet
            ws['B' + str(k)] = y
            k = k + 1
        for z in getRequest.resultsList:
            resultString = str(z)
            if resultString == '200' :
                ws['C' + str(l)] = 'ok'
                l = l+1
            else:
                ws['C' + str(l)] = 'check status'
                l = l+1
                
        resultsWorkBook.save(fileName)                      # save the excel spreadsheet to downloads folder
            
        
def main():
    openFile.getSheet()
    getRequest.returnRequestStatus(openFile.dataList)
    resultFile.createFile()
    
if __name__ =='__main__':
    main()      
